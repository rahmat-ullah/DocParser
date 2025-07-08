"""
XLSX Parser using openpyxl.
Extracts tables and text content from Excel files.
"""

from typing import Optional, AsyncGenerator
from pathlib import Path
from openpyxl import load_workbook

from .base_parser import BaseParser, ParseError
from .ast_models import DocumentAST, TableBlock, ParseProgress


class XLSXParser(BaseParser):
    """Parser for XLSX documents using openpyxl."""

    def supports_file(self, file_path: Path) -> bool:
        """Check if file is a XLSX."""
        return file_path.suffix.lower() in ['.xlsx', '.xls']

    async def parse(
        self, file_path: Path, progress_callback: Optional[AsyncGenerator[ParseProgress, None]] = None
    ) -> DocumentAST:
        """Parse XLSX document and extract content."""
        try:
            await self._emit_progress(progress_callback, "initialization", 0.0, "Opening XLSX document")

            workbook = load_workbook(file_path, data_only=True)
            ast = DocumentAST(metadata={"format": "XLSX", "sheets": workbook.sheetnames})

            total_sheets = len(workbook.sheetnames)
            
            for i, sheet_name in enumerate(workbook.sheetnames):
                await self._emit_progress(
                    progress_callback, 
                    "parsing_sheets", 
                    i / total_sheets, 
                    f"Processing sheet: {sheet_name}"
                )
                
                worksheet = workbook[sheet_name]
                
                # Extract table data from each worksheet
                if worksheet.max_row > 0:
                    # Get headers from first row
                    headers = []
                    for cell in worksheet[1]:
                        headers.append(str(cell.value) if cell.value is not None else "")
                    
                    # Get data rows
                    rows = []
                    for row in worksheet.iter_rows(min_row=2, values_only=True):
                        row_data = [str(cell) if cell is not None else "" for cell in row]
                        if any(row_data):  # Skip empty rows
                            rows.append(row_data)
                    
                    if headers and rows:
                        table_block = TableBlock(
                            headers=headers,
                            rows=rows,
                            caption=f"Sheet: {sheet_name}"
                        )
                        ast.tables.append(table_block)

                    # Convert to Markdown
                    markdown_table = "| " + " | ".join(headers) + " |\n"
                    markdown_table += "|" + "---|" * len(headers) + "\n"
                    for row in rows:
                        markdown_table += "| " + " | ".join(row) + " |\n"
                    ast.metadata[f"markdown_{sheet_name}"] = markdown_table

            await self._emit_progress(progress_callback, "completion", 1.0, "XLSX parsing completed")
            return ast

        except Exception as e:
            raise ParseError(f"Failed to parse XLSX: {str(e)}", file_path, e)
