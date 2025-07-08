import pytest
import asyncio
from pathlib import Path
from openpyxl import Workbook

from app.parsers.xlsx_parser import XLSXParser

@pytest.fixture
def xlsx_parser():
    return XLSXParser()

@pytest.fixture
def xlsx_file(tmp_path):
    # Create a simple XLSX file with content
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Sheet"
    
    # Add headers
    ws['A1'] = 'Name'
    ws['B1'] = 'Age'
    ws['C1'] = 'City'
    
    # Add data
    ws['A2'] = 'John Doe'
    ws['B2'] = 30
    ws['C2'] = 'New York'
    
    ws['A3'] = 'Jane Smith'
    ws['B3'] = 25
    ws['C3'] = 'Los Angeles'
    
    path = tmp_path / "test.xlsx"
    wb.save(str(path))
    return path

@pytest.mark.asyncio
async def test_xlsx_parser_supports_file(xlsx_parser, xlsx_file):
    assert xlsx_parser.supports_file(xlsx_file) == True
    
@pytest.mark.asyncio
async def test_xlsx_parser_parse(xlsx_parser, xlsx_file):
    result = await xlsx_parser.parse_to_dict(xlsx_file)
    assert isinstance(result, dict)
    assert "metadata" in result
    assert result["metadata"]["format"] == "XLSX"
    assert "tables" in result
    assert len(result["tables"]) > 0
    
    # Check if table was extracted
    table = result["tables"][0]
    assert "headers" in table
    assert len(table["headers"]) == 3
    assert table["headers"] == ['Name', 'Age', 'City']
    assert "rows" in table
    assert len(table["rows"]) == 2
    
    # Check if markdown was generated
    assert "markdown_Test Sheet" in result["metadata"]
    markdown = result["metadata"]["markdown_Test Sheet"]
    assert "| Name | Age | City |" in markdown
